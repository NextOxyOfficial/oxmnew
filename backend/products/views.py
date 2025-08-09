import csv
import io
import json

import openpyxl
import xlrd
from django.db import transaction
from django.db.models import Count, Q, Sum
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters, status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from .models import Product, ProductPhoto, ProductStockMovement, ProductVariant
from .serializers import (
    ProductCreateSerializer,
    ProductDetailSerializer,
    ProductListSerializer,
    ProductPhotoSerializer,
    ProductStockMovementSerializer,
    ProductVariantSerializer,
)


class ProductViewSet(viewsets.ModelViewSet):
    """ViewSet for managing products"""

    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = ["category", "supplier", "has_variants", "is_active"]
    search_fields = [
        "name",
        "product_code",
        "details",
        "location",
        "category__name",
        "supplier__name",
    ]
    ordering_fields = ["name", "created_at", "updated_at", "buy_price", "sell_price"]
    ordering = ["-created_at"]

    def get_queryset(self):
        """Return products for the authenticated user"""
        return (
            Product.objects.filter(user=self.request.user)
            .select_related("category", "supplier")
            .prefetch_related("variants", "photos")
        )

    def get_serializer_class(self):
        """Return appropriate serializer based on action"""
        if self.action == "list":
            return ProductListSerializer
        elif self.action == "create":
            return ProductCreateSerializer
        else:
            return ProductDetailSerializer

    def create(self, request, *args, **kwargs):
        """Create a new product with custom handling for FormData"""
        data = request.data.copy()

        print("=== BACKEND CREATE PRODUCT DEBUG ===")
        print("Original request.data keys:", list(request.data.keys()))
        print("Original request.FILES keys:", list(request.FILES.keys()))

        # Handle photos from FormData
        photos = []

        # Check if photos are sent as a list with the same key
        if "photos" in request.FILES:
            photos_files = request.FILES.getlist("photos")
            photos.extend(photos_files)
            print(f"Found {len(photos_files)} photos with key 'photos'")

        # Also check for photos with different keys (e.g., photos[0], photos[1], etc.)
        for key in request.FILES:
            if key.startswith("photos") and key != "photos":
                photos.append(request.FILES[key])
                print(f"Found photo with key '{key}'")

        print(f"Total photos found: {len(photos)}")

        # Handle colorSizeVariants JSON string
        if "colorSizeVariants" in data and isinstance(data["colorSizeVariants"], str):
            try:
                parsed_variants = json.loads(data["colorSizeVariants"])

                # Handle case where the variants might be nested in another array
                # This can happen if the data gets processed multiple times
                if (
                    isinstance(parsed_variants, list)
                    and len(parsed_variants) == 1
                    and isinstance(parsed_variants[0], list)
                ):
                    print("Detected nested array in colorSizeVariants, flattening...")
                    parsed_variants = parsed_variants[0]

                data["colorSizeVariants"] = parsed_variants
                print("Parsed colorSizeVariants:", parsed_variants)
                print("Type after parsing:", type(parsed_variants))
            except json.JSONDecodeError as e:
                print("JSON decode error:", str(e))
                print("Raw colorSizeVariants data:", repr(data["colorSizeVariants"]))
                return Response(
                    {"error": f"Invalid colorSizeVariants JSON: {str(e)}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        # Convert string boolean values
        if "hasVariants" in data:
            data["hasVariants"] = data["hasVariants"] in ["true", "True", True]
            print("Converted hasVariants:", data["hasVariants"])
        elif "has_variants" in data:
            data["has_variants"] = data["has_variants"] in ["true", "True", True]
            print("Converted has_variants:", data["has_variants"])

        # Convert numeric fields - handle both snake_case and camelCase
        numeric_fields = ["buyPrice", "sellPrice", "stock", "buy_price", "sell_price"]
        for field in numeric_fields:
            if field in data and data[field]:
                try:
                    if field == "stock":
                        data[field] = int(data[field])
                    else:
                        data[field] = float(data[field])
                    print(f"Converted {field}:", data[field])
                except (ValueError, TypeError):
                    data[field] = 0
                    print(f"Failed to convert {field}, set to 0")

        print(
            "Final data for serializer:",
            {k: v for k, v in data.items() if k != "photos"},
        )
        print("Photos in data:", len(photos))

        # Debug colorSizeVariants specifically
        if "colorSizeVariants" in data:
            print("=== colorSizeVariants DEBUG ===")
            print("Type:", type(data["colorSizeVariants"]))
            print("Value:", data["colorSizeVariants"])
            print("Is list?", isinstance(data["colorSizeVariants"], list))
            if (
                isinstance(data["colorSizeVariants"], list)
                and data["colorSizeVariants"]
            ):
                print("First element type:", type(data["colorSizeVariants"][0]))
                print("First element:", data["colorSizeVariants"][0])
        else:
            print("colorSizeVariants not in data")

        # Don't add photos to data - handle them separately
        # Remove any photos from data to avoid validation issues
        if "photos" in data:
            del data["photos"]

        # Prepare clean data for serializer to avoid QueryDict nesting issues
        # Only include fields that are supported by the serializer
        serializer_class = self.get_serializer_class()
        valid_fields = set(serializer_class.Meta.fields)

        # Map snake_case field names to camelCase for the serializer
        field_mapping = {
            "buy_price": "buyPrice",
            "sell_price": "sellPrice",
            "product_code": "productCode",
            "has_variants": "hasVariants",
        }

        serializer_data = {}
        for key, value in data.items():
            # Check if it's a direct valid field or needs mapping
            mapped_key = field_mapping.get(key, key)

            if (
                mapped_key in valid_fields or key == "colorSizeVariants"
            ):  # colorSizeVariants is write_only
                if key == "colorSizeVariants" and isinstance(value, list):
                    # Ensure we pass the list directly, not nested
                    if value and isinstance(value[0], list):
                        print(
                            "WARNING: Flattening nested colorSizeVariants before serializer"
                        )
                        serializer_data[key] = value[0]
                    else:
                        serializer_data[key] = value
                else:
                    # Use the mapped key for the serializer
                    serializer_data[mapped_key] = value
            else:
                print(f"Skipping unknown field: {key}")

        print("=== SERIALIZER INPUT DATA ===")
        print("serializer_data:", serializer_data)
        if "colorSizeVariants" in serializer_data:
            print("colorSizeVariants type:", type(serializer_data["colorSizeVariants"]))
            print("colorSizeVariants value:", serializer_data["colorSizeVariants"])

        # Create serializer with clean data
        serializer = self.get_serializer(data=serializer_data)
        print("Serializer data before validation:", serializer.initial_data)

        if serializer.is_valid():
            print("Serializer is valid, proceeding to save")
            # Save the product first
            product = serializer.save()

            # Now handle photos separately
            if photos:
                print(f"Creating {len(photos)} ProductPhoto objects")
                for i, photo in enumerate(photos):
                    ProductPhoto.objects.create(product=product, image=photo, order=i)

            response_serializer = ProductDetailSerializer(
                product, context={"request": request}
            )
            return Response(response_serializer.data, status=status.HTTP_201_CREATED)
        else:
            print("Serializer validation errors:", serializer.errors)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=["post"])
    def add_variant(self, request, pk=None):
        """Add a new variant to an existing product"""
        product = self.get_object()

        if not product.has_variants:
            return Response(
                {"error": "This product does not support variants"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        serializer = ProductVariantSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(product=product)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(
        detail=True,
        methods=["patch", "delete"],
        url_path="variants/(?P<variant_id>[^/.]+)",
    )
    def manage_variant(self, request, pk=None, variant_id=None):
        """Update or delete a specific variant"""
        product = self.get_object()

        try:
            variant = product.variants.get(id=variant_id)
        except ProductVariant.DoesNotExist:
            return Response(
                {"error": "Variant not found"}, status=status.HTTP_404_NOT_FOUND
            )

        if request.method == "PATCH":
            serializer = ProductVariantSerializer(
                variant, data=request.data, partial=True
            )
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        elif request.method == "DELETE":
            variant.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=["post"])
    def add_photos(self, request, pk=None):
        """Add photos to a product"""
        product = self.get_object()
        photos = request.FILES.getlist("photos")

        if not photos:
            return Response(
                {"error": "No photos provided"}, status=status.HTTP_400_BAD_REQUEST
            )

        created_photos = []
        for i, photo in enumerate(photos):
            # Get the current max order
            max_order = (
                product.photos.aggregate(max_order=Sum("order"))["max_order"] or 0
            )

            photo_obj = ProductPhoto.objects.create(
                product=product, image=photo, order=max_order + i + 1
            )
            created_photos.append(photo_obj)

        serializer = ProductPhotoSerializer(created_photos, many=True)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=["delete"], url_path="photos/(?P<photo_id>[^/.]+)")
    def delete_photo(self, request, pk=None, photo_id=None):
        """Delete a specific photo"""
        product = self.get_object()

        try:
            photo = product.photos.get(id=photo_id)
        except ProductPhoto.DoesNotExist:
            return Response(
                {"error": "Photo not found"}, status=status.HTTP_404_NOT_FOUND
            )

        photo.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=["post"])
    def adjust_stock(self, request, pk=None):
        """Adjust product stock"""
        product = self.get_object()
        variant_id = request.data.get("variant_id")
        quantity = request.data.get("quantity")
        reason = request.data.get("reason", "Manual adjustment")
        notes = request.data.get("notes", "")
        buy_price = request.data.get("buy_price")
        update_average_price = request.data.get("update_average_price", False)
        new_average_buy_price = request.data.get("new_average_buy_price")

        if quantity is None:
            return Response(
                {"error": "Quantity is required"}, status=status.HTTP_400_BAD_REQUEST
            )

        try:
            quantity = int(quantity)
        except (ValueError, TypeError):
            return Response(
                {"error": "Invalid quantity"}, status=status.HTTP_400_BAD_REQUEST
            )

        # Parse buy_price if provided
        if buy_price is not None:
            try:
                buy_price = float(buy_price)
            except (ValueError, TypeError):
                return Response(
                    {"error": "Invalid buy price"}, status=status.HTTP_400_BAD_REQUEST
                )

        # Parse new_average_buy_price if provided
        if new_average_buy_price is not None:
            try:
                new_average_buy_price = float(new_average_buy_price)
            except (ValueError, TypeError):
                return Response(
                    {"error": "Invalid average buy price"}, status=status.HTTP_400_BAD_REQUEST
                )

        with transaction.atomic():
            if variant_id:
                try:
                    variant = product.variants.get(id=variant_id)
                except ProductVariant.DoesNotExist:
                    return Response(
                        {"error": "Variant not found"}, status=status.HTTP_404_NOT_FOUND
                    )

                previous_stock = variant.stock
                new_stock = max(0, previous_stock + quantity)
                variant.stock = new_stock
                variant.save()

                # Record stock movement
                movement_data = {
                    "product": product,
                    "variant": variant,
                    "user": request.user,
                    "movement_type": "adjustment",
                    "quantity": quantity,
                    "previous_stock": previous_stock,
                    "new_stock": new_stock,
                    "reason": reason,
                    "notes": notes,
                }
                
                # Add cost information if buy_price is provided
                if buy_price is not None:
                    movement_data["cost_per_unit"] = buy_price
                    movement_data["total_cost"] = buy_price * abs(quantity)

                ProductStockMovement.objects.create(**movement_data)

                return Response(
                    {
                        "message": "Stock adjusted successfully",
                        "previous_stock": previous_stock,
                        "new_stock": new_stock,
                        "variant_id": variant.id,
                    }
                )
            else:
                previous_stock = product.stock
                new_stock = max(0, previous_stock + quantity)
                product.stock = new_stock
                
                # Update average buy price if requested and provided
                if update_average_price and new_average_buy_price is not None:
                    product.buy_price = new_average_buy_price
                
                product.save()

                # Record stock movement
                movement_data = {
                    "product": product,
                    "user": request.user,
                    "movement_type": "adjustment",
                    "quantity": quantity,
                    "previous_stock": previous_stock,
                    "new_stock": new_stock,
                    "reason": reason,
                    "notes": notes,
                }
                
                # Add cost information if buy_price is provided
                if buy_price is not None:
                    movement_data["cost_per_unit"] = buy_price
                    movement_data["total_cost"] = buy_price * abs(quantity)

                ProductStockMovement.objects.create(**movement_data)

                response_data = {
                    "message": "Stock adjusted successfully",
                    "previous_stock": previous_stock,
                    "new_stock": new_stock,
                }
                
                if update_average_price and new_average_buy_price is not None:
                    response_data["new_average_buy_price"] = new_average_buy_price
                    response_data["buy_price_updated"] = True

                return Response(response_data)

    def _extract_data_from_file(self, file):
        """Extract data from CSV, XLSX, or XLS file"""
        file.seek(0)  # Reset file pointer

        if file.name.endswith(".csv"):
            # Handle CSV file
            decoded_file = file.read().decode("utf-8")
            lines = decoded_file.splitlines()
            csv_reader = csv.DictReader(lines)
            return list(csv_reader)

        elif file.name.endswith(".xlsx"):
            # Handle Excel XLSX file
            workbook = openpyxl.load_workbook(file, read_only=True)
            worksheet = workbook.active

            # Get headers from first row
            headers = []
            for cell in worksheet[1]:
                if cell.value is not None:
                    headers.append(str(cell.value).lower().strip())
                else:
                    break

            # Extract data rows
            data = []
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):  # Skip empty rows
                    row_data = {}
                    for i, value in enumerate(row):
                        if i < len(headers) and headers[i]:
                            row_data[headers[i]] = (
                                str(value) if value is not None else ""
                            )
                    if row_data:  # Only add non-empty rows
                        data.append(row_data)

            workbook.close()
            return data

        elif file.name.endswith(".xls"):
            # Handle Excel XLS file
            file_contents = file.read()
            workbook = xlrd.open_workbook(file_contents=file_contents)
            worksheet = workbook.sheet_by_index(0)

            # Get headers from first row
            headers = []
            for col in range(worksheet.ncols):
                cell_value = worksheet.cell_value(0, col)
                if cell_value:
                    headers.append(str(cell_value).lower().strip())
                else:
                    break

            # Extract data rows
            data = []
            for row in range(1, worksheet.nrows):
                row_data = {}
                for col in range(min(len(headers), worksheet.ncols)):
                    if headers[col]:
                        cell_value = worksheet.cell_value(row, col)
                        row_data[headers[col]] = str(cell_value) if cell_value else ""
                if any(value for value in row_data.values()):  # Only add non-empty rows
                    data.append(row_data)

            return data

        else:
            raise ValueError(
                "Unsupported file format. Please use CSV, XLSX, or XLS files."
            )

    @action(detail=False, methods=["post"])
    def upload_csv(self, request):
        """Upload products from CSV, XLSX, or XLS file"""
        if "csv_file" not in request.FILES:
            return Response(
                {"error": "No file provided"}, status=status.HTTP_400_BAD_REQUEST
            )

        upload_file = request.FILES["csv_file"]

        # Check file extension
        allowed_extensions = [".csv", ".xlsx", ".xls"]
        file_extension = None
        for ext in allowed_extensions:
            if upload_file.name.lower().endswith(ext):
                file_extension = ext
                break

        if not file_extension:
            return Response(
                {
                    "error": "File must be a CSV (.csv), Excel (.xlsx), or Excel (.xls) file"
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            # Extract data from file
            rows_data = self._extract_data_from_file(upload_file)

            products_created = 0
            products_errors = []
            successful_rows = []

            with transaction.atomic():
                for row_num, row in enumerate(
                    rows_data, start=2
                ):  # Start at 2 to account for header
                    try:
                        # Clean and validate data
                        name = row.get("name", "").strip()
                        if not name:
                            products_errors.append(
                                f"Row {row_num}: Product name is required (cannot be empty)"
                            )
                            continue

                        # Check for existing product with same name for this user
                        if Product.objects.filter(
                            name=name, user=request.user
                        ).exists():
                            products_errors.append(
                                f"Row {row_num}: Product '{name}' already exists"
                            )
                            continue

                        # Parse basic fields
                        product_data = {
                            "name": name,
                            "product_code": row.get("product_code", "").strip() or None,
                            "location": row.get("location", "").strip() or None,
                            "details": row.get("details", "").strip() or "",
                            "user": request.user,
                            "has_variants": False,  # Default for CSV upload
                        }

                        # Parse prices with validation
                        buy_price_str = str(row.get("buy_price", "")).strip()
                        sell_price_str = str(row.get("sell_price", "")).strip()
                        stock_str = str(row.get("stock", "")).strip()

                        # Check for empty required fields
                        missing_fields = []
                        if not buy_price_str or buy_price_str.lower() in [
                            "",
                            "none",
                            "null",
                        ]:
                            missing_fields.append("buy_price")
                        if not sell_price_str or sell_price_str.lower() in [
                            "",
                            "none",
                            "null",
                        ]:
                            missing_fields.append("sell_price")
                        if not stock_str or stock_str.lower() in ["", "none", "null"]:
                            missing_fields.append("stock")

                        if missing_fields:
                            missing_fields_text = ", ".join(missing_fields)
                            products_errors.append(
                                f"Row {row_num}: Product '{name}' is missing required field(s): {missing_fields_text}"
                            )
                            continue

                        try:
                            buy_price = float(buy_price_str)
                            sell_price = float(sell_price_str)

                            if buy_price <= 0:
                                products_errors.append(
                                    f"Row {row_num}: Product '{name}' - Buy price must be greater than 0"
                                )
                                continue
                            if sell_price <= 0:
                                products_errors.append(
                                    f"Row {row_num}: Product '{name}' - Sell price must be greater than 0"
                                )
                                continue
                            if sell_price < buy_price:
                                products_errors.append(
                                    f"Row {row_num}: Product '{name}' - Sell price must be >= buy price"
                                )
                                continue

                            product_data["buy_price"] = buy_price
                            product_data["sell_price"] = sell_price
                        except (ValueError, TypeError):
                            products_errors.append(
                                f"Row {row_num}: Product '{name}' - Invalid price values (buy_price: '{buy_price_str}', sell_price: '{sell_price_str}')"
                            )
                            continue

                        # Parse stock
                        try:
                            stock = int(
                                float(stock_str)
                            )  # Convert via float first to handle decimal inputs
                            if stock <= 0:
                                products_errors.append(
                                    f"Row {row_num}: Product '{name}' - Stock quantity must be greater than 0"
                                )
                                continue
                            product_data["stock"] = stock
                        except (ValueError, TypeError):
                            products_errors.append(
                                f"Row {row_num}: Product '{name}' - Invalid stock value ('{stock_str}')"
                            )
                            continue

                        # Handle category
                        category_name = row.get("category", "").strip()
                        if category_name:
                            try:
                                from core.models import Category

                                # First try to get existing category by name (global)
                                try:
                                    category = Category.objects.get(name=category_name)
                                except Category.DoesNotExist:
                                    # If category doesn't exist, create a new one
                                    category = Category.objects.create(
                                        name=category_name,
                                        user=request.user,
                                        is_active=True,
                                    )
                                product_data["category"] = category
                            except Exception as e:
                                products_errors.append(
                                    f"Row {row_num}: Product '{name}' - Error with category '{category_name}': {str(e)}"
                                )
                                continue

                        # Handle supplier
                        supplier_name = row.get("supplier", "").strip()
                        if supplier_name:
                            try:
                                from suppliers.models import Supplier

                                # First try to get existing supplier by name and user
                                try:
                                    supplier = Supplier.objects.get(
                                        name=supplier_name, user=request.user
                                    )
                                except Supplier.DoesNotExist:
                                    # If supplier doesn't exist for this user, create a new one
                                    supplier = Supplier.objects.create(
                                        name=supplier_name,
                                        user=request.user,
                                        is_active=True,
                                    )
                                product_data["supplier"] = supplier
                            except Exception as e:
                                products_errors.append(
                                    f"Row {row_num}: Product '{name}' - Error with supplier '{supplier_name}': {str(e)}"
                                )
                                continue

                        # Create the product
                        product = Product.objects.create(**product_data)
                        products_created += 1
                        successful_rows.append({"row": row_num, "name": product.name})

                    except Exception as e:
                        # Try to include product name in error if available
                        error_msg = f"Row {row_num}: "
                        if "name" in locals() and name:
                            error_msg += f"Product '{name}' - "
                        error_msg += str(e)
                        products_errors.append(error_msg)
                        continue

            return Response(
                {
                    "success": True,
                    "products_created": products_created,
                    "successful_rows": successful_rows,
                    "errors": products_errors,
                    "message": f"Successfully created {products_created} products",
                },
                status=status.HTTP_201_CREATED,
            )

        except Exception as e:
            return Response(
                {"error": f"Error processing file: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=["get"])
    def download_csv_template(self, request):
        """Download CSV template for product upload"""
        from django.http import HttpResponse

        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="products_template.csv"'

        writer = csv.writer(response)
        writer.writerow(
            [
                "name",
                "product_code",
                "category",
                "supplier",
                "location",
                "details",
                "buy_price",
                "sell_price",
                "stock",
            ]
        )

        # Add sample data
        writer.writerow(
            [
                "Sample Product 1",
                "SP001",
                "Electronics",
                "Sample Supplier",
                "Warehouse A",
                "Sample product description",
                "50.00",
                "75.00",
                "100",
            ]
        )
        writer.writerow(
            [
                "Sample Product 2",
                "",  # Empty product_code to show it's optional
                "Clothing",
                "Another Supplier",
                "Store Room B",
                "Another sample product",
                "25.00",
                "40.00",
                "50",
            ]
        )

        return response

    @action(detail=False, methods=["get"])
    def download_excel_template(self, request):
        """Download Excel template for product upload"""
        from django.http import HttpResponse

        # Create a new workbook and add a worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Products"

        # Define headers
        headers = [
            "name",
            "product_code",
            "category",
            "supplier",
            "location",
            "details",
            "buy_price",
            "sell_price",
            "stock",
        ]

        # Add headers to first row
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)

        # Add sample data
        sample_data = [
            [
                "Sample Product 1",
                "SP001",
                "Electronics",
                "Sample Supplier",
                "Warehouse A",
                "Sample product description",
                50.00,
                75.00,
                100,
            ],
            [
                "Sample Product 2",
                "SP002",
                "Clothing",
                "Another Supplier",
                "Store Room B",
                "Another sample product",
                25.00,
                40.00,
                50,
            ],
            [
                "Sample Product 3",
                "",  # Empty product_code to show it's optional
                "Home & Garden",
                "Third Supplier",
                "Storage C",
                "Third sample product with longer description",
                15.50,
                25.99,
                75,
            ],
        ]

        # Add sample data rows
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Save workbook to memory
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        # Create response
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            'attachment; filename="products_template.xlsx"'
        )

        return response

    @action(detail=False, methods=["get"])
    def statistics(self, request):
        """Get product statistics"""
        queryset = self.get_queryset()

        total_products = queryset.count()
        active_products = queryset.filter(is_active=True).count()
        products_with_variants = queryset.filter(has_variants=True).count()
        total_stock = sum(product.total_stock for product in queryset)

        # Calculate total inventory value
        total_value = 0
        for product in queryset:
            if product.has_variants:
                for variant in product.variants.all():
                    total_value += variant.stock * variant.buy_price
            else:
                total_value += product.stock * product.buy_price

        # Low stock products (less than 10 items)
        low_stock_products = []
        for product in queryset:
            if product.has_variants:
                for variant in product.variants.all():
                    if variant.stock < 10:
                        low_stock_products.append(
                            {
                                "product": product.name,
                                "variant": str(variant),
                                "stock": variant.stock,
                            }
                        )
            else:
                if product.stock < 10:
                    low_stock_products.append(
                        {
                            "product": product.name,
                            "variant": None,
                            "stock": product.stock,
                        }
                    )

        return Response(
            {
                "total_products": total_products,
                "active_products": active_products,
                "products_with_variants": products_with_variants,
                "total_stock": total_stock,
                "total_inventory_value": total_value,
                "low_stock_count": len(low_stock_products),
                "low_stock_products": low_stock_products[:10],  # Limit to first 10
            }
        )

    @action(detail=False, methods=["get"])
    def stats(self, request):
        """Get product statistics"""
        queryset = self.get_queryset()

        total_products = queryset.count()
        active_products = queryset.filter(is_active=True).count()
        low_stock_products = queryset.filter(stock__lte=10).count()
        out_of_stock_products = queryset.filter(stock=0).count()

        # Calculate total inventory value
        total_buy_value = sum(product.total_buy_price for product in queryset)
        total_sell_value = sum(product.total_sell_price for product in queryset)

        # Get top categories
        category_stats = (
            queryset.values("category__name")
            .annotate(count=Count("id"))
            .order_by("-count")[:5]
        )

        return Response(
            {
                "total_products": total_products,
                "active_products": active_products,
                "low_stock_products": low_stock_products,
                "out_of_stock_products": out_of_stock_products,
                "total_buy_value": float(total_buy_value),
                "total_sell_value": float(total_sell_value),
                "top_categories": list(category_stats),
            }
        )

    @action(detail=False, methods=["get"])
    def search(self, request):
        """Advanced search endpoint for products"""
        query = request.query_params.get("q", "").strip()

        if not query or len(query) < 2:
            return Response({"results": [], "total": 0})

        queryset = self.get_queryset()

        # Search in multiple fields
        search_results = queryset.filter(
            Q(name__icontains=query)
            | Q(details__icontains=query)
            | Q(location__icontains=query)
            | Q(category__name__icontains=query)
            | Q(supplier__name__icontains=query)
        ).distinct()

        # Format results for frontend
        results = []
        for product in search_results[:10]:  # Limit to 10 results
            stock_level = (
                "HIGH_STOCK"
                if product.stock > 20
                else "MEDIUM_STOCK"
                if product.stock > 10
                else "LOW_STOCK"
            )

            results.append(
                {
                    "id": str(product.id),
                    "type": "product",
                    "title": product.name,
                    "subtitle": f"{product.category.name if product.category else 'Uncategorized'} • ${product.sell_price} • {product.stock} in stock",
                    "href": f"/dashboard/products/{product.id}",
                    "badge": stock_level,
                }
            )

        return Response(
            {"results": results, "total": search_results.count(), "query": query}
        )


class ProductStockMovementViewSet(viewsets.ModelViewSet):
    """ViewSet for viewing and managing stock movements"""

    serializer_class = ProductStockMovementSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ["product", "variant", "movement_type"]
    ordering_fields = ["created_at"]
    ordering = ["-created_at"]

    def get_queryset(self):
        """Return stock movements for the authenticated user"""
        return ProductStockMovement.objects.filter(
            user=self.request.user
        ).select_related("product", "variant")

    def destroy(self, request, *args, **kwargs):
        """Delete a stock movement and recalculate product buy price"""
        movement = self.get_object()
        product = movement.product
        variant = movement.variant
        
        # Store movement data before deletion
        movement_quantity = movement.quantity
        movement_type = movement.movement_type
        
        # Delete the movement first
        self.perform_destroy(movement)
        
        # Reverse the stock change
        if variant:
            # For variant stock
            if movement_type in ['in', 'adjustment'] and movement_quantity > 0:
                # Was stock addition, so subtract it back
                variant.stock = max(0, variant.stock - movement_quantity)
            elif movement_type in ['out', 'sale'] and movement_quantity < 0:
                # Was stock reduction, so add it back
                variant.stock += abs(movement_quantity)
            variant.save()
        else:
            # For main product stock
            if movement_type in ['in', 'adjustment'] and movement_quantity > 0:
                # Was stock addition, so subtract it back
                product.stock = max(0, product.stock - movement_quantity)
            elif movement_type in ['out', 'sale'] and movement_quantity < 0:
                # Was stock reduction, so add it back
                product.stock += abs(movement_quantity)
        
        # Recalculate the weighted average buy price from remaining movements
        # Get all remaining stock movements for this product that added stock and had cost
        remaining_movements = ProductStockMovement.objects.filter(
            product=product,
            quantity__gt=0,  # Only stock additions
            cost_per_unit__isnull=False,
            cost_per_unit__gt=0
        ).order_by('created_at')
        
        if remaining_movements.exists():
            # Calculate weighted average from remaining movements
            total_cost = 0
            total_quantity = 0
            
            for mov in remaining_movements:
                cost = float(mov.cost_per_unit)
                qty = mov.quantity
                total_cost += cost * qty
                total_quantity += qty
            
            if total_quantity > 0:
                new_avg_price = total_cost / total_quantity
                product.buy_price = round(new_avg_price, 2)
            else:
                # If no valid movements remain, keep current price
                pass
        else:
            # If no remaining cost movements, you might want to:
            # Option 1: Keep current price (safer)
            # Option 2: Reset to 0 (commented below)
            # product.buy_price = 0
            pass
        
        product.save()
        
        return Response({
            'message': 'Stock movement deleted successfully',
            'new_buy_price': float(product.buy_price) if product.buy_price else 0
        }, status=status.HTTP_200_OK)
