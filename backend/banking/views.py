from core.scoping import HasPermission, owner_for
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation

from django.db import transaction as db_transaction
from django.db.models import Q
from django.utils import timezone
from django_filters.rest_framework import DjangoFilterBackend
from employees.models import Employee
from rest_framework import filters, generics, status, viewsets
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.exceptions import PermissionDenied
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from core.uploads import document_error
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination
from django.db.models import Sum

from .models import (
    BankAccount,
    BankingPlan,
    Loan,
    LoanPayment,
    RecurringCost,
    RecurringCostPayment,
    Transaction,
    UserBankingPlan,
)
from .serializers import (
    LoanPaymentSerializer,
    LoanSerializer,
    RecurringCostPaymentSerializer,
    RecurringCostSerializer,
    BankAccountSerializer,
    BankingPlanSerializer,
    TransactionCreateSerializer,
    TransactionSerializer,
)


class TransactionPagination(PageNumberPagination):
    """Custom pagination for transactions"""
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 100


def has_active_banking_plan(user):
    """Check if user has a pro subscription plan"""
    try:
        from subscription.models import UserSubscription
        user_subscription = UserSubscription.objects.get(user=user, active=True)
        return user_subscription.plan.name == 'pro'
    except UserSubscription.DoesNotExist:
        return False


class BankAccountViewSet(viewsets.ModelViewSet):
    # Staff logins are held to these; owners are unrestricted.
    required_permissions = {
        "GET": "banking.view",
        "POST": "banking.transact",
        "PUT": "banking.transact",
        "PATCH": "banking.transact",
        "DELETE": "banking.transact",
    }

    serializer_class = BankAccountSerializer
    permission_classes = [IsAuthenticated, HasPermission]
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = ["is_active", "owner"]
    search_fields = ["name", "owner__username", "owner__first_name", "owner__last_name"]
    ordering_fields = ["name", "balance", "created_at"]
    ordering = ["-created_at"]

    def get_queryset(self):
        """Return accounts based on user permissions"""
        user = owner_for(self.request)
        # Ensure user has a Main account
        self.ensure_main_account(user)

        # If user is staff/admin, they can see all accounts
        if user.is_staff or user.is_superuser:
            queryset = BankAccount.objects.filter(is_active=True)
        else:
            # Regular users can only see their own accounts
            queryset = BankAccount.objects.filter(owner=user, is_active=True)

        # Order by: Main account first, then by creation date (newest first)
        return queryset.extra(
            select={"is_main": "CASE WHEN name = 'Main' THEN 0 ELSE 1 END"}
        ).order_by("is_main", "-created_at")

    def ensure_main_account(self, user):
        """Ensure user has a Main account"""
        if not BankAccount.objects.filter(owner=user, name="Main").exists():
            BankAccount.objects.create(
                name="Main",
                owner=user,
                balance=0.00,
                is_active=True,
                activation_fee=0.00,
                is_activated=True,
            )

    def perform_create(self, serializer):
        """Set the current user as the account owner and check account limits"""
        import logging
        logger = logging.getLogger(__name__)
        
        user = owner_for(self.request)
        logger.info(f"🏦 User {user.username} attempting to create bank account")
        
        # Get existing active accounts count
        existing_accounts_count = BankAccount.objects.filter(
            owner=user, is_active=True
        ).count()
        logger.info(f"🏦 User {user.username} has {existing_accounts_count} existing accounts")
        
        # Check if user has pro subscription
        has_pro = has_active_banking_plan(user)
        logger.info(f"🏦 User {user.username} has pro subscription: {has_pro}")
        
        if has_pro:
            # Pro users can have up to 15 accounts
            if existing_accounts_count >= 15:
                from rest_framework.exceptions import PermissionDenied
                logger.warning(f"🏦 User {user.username} exceeded pro account limit: {existing_accounts_count}/15")
                raise PermissionDenied(
                    f"You have reached the maximum limit of 15 accounts. "
                    f"Please delete an existing account before creating a new one."
                )
        else:
            # Free users can only have 1 account
            if existing_accounts_count >= 1:
                from rest_framework.exceptions import PermissionDenied
                logger.warning(f"🏦 User {user.username} exceeded free account limit: {existing_accounts_count}/1")
                raise PermissionDenied(
                    "You need to upgrade to Pro to create additional accounts. "
                    "Free users can only have 1 account."
                )
        
        # Set balance from the payload (frontend sends 'balance' field)
        balance = serializer.validated_data.get('balance', 0)
        logger.info(f"🏦 Creating account with balance: {balance}")
        
        try:
            account = serializer.save(owner=user, balance=balance)
            logger.info(f"🏦 Successfully created account: {account.id} - {account.name}")
        except Exception as e:
            logger.error(f"🏦 Failed to create account: {str(e)}")
            raise

    @action(detail=False, methods=["get"])
    def my_accounts(self, request):
        """Get current user's accounts only"""
        accounts = (
            BankAccount.objects.filter(owner=owner_for(request), is_active=True)
            .extra(select={"is_main": "CASE WHEN name = 'Main' THEN 0 ELSE 1 END"})
            .order_by("is_main", "-created_at")
        )
        serializer = self.get_serializer(accounts, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=["get"])
    def transactions(self, request, pk=None):
        """Get all transactions for a specific account with pagination"""
        try:
            account = BankAccount.objects.get(id=pk, is_active=True)

            if not request.user.is_staff and not request.user.is_superuser:
                if account.owner != request.user:
                    return Response(
                        {"error": "You don't have permission to access this account."},
                        status=status.HTTP_403_FORBIDDEN
                    )

        except BankAccount.DoesNotExist:
            return Response(
                {"error": f"Bank account with ID {pk} not found."},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"error": f"Error accessing account: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        transactions = account.transactions.all().order_by('-date', '-updated_at')

        # Apply filtering
        transaction_type = request.query_params.get("type")
        status_filter = request.query_params.get("status")
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")
        search = request.query_params.get("search")
        verified_by = request.query_params.get("verified_by")

        if transaction_type and transaction_type != "all":
            transactions = transactions.filter(type=transaction_type)
        if status_filter and status_filter != "all":
            transactions = transactions.filter(status=status_filter)
        if date_from:
            transactions = transactions.filter(date__gte=date_from)
        if date_to:
            from datetime import datetime, time
            try:
                end_date = datetime.strptime(date_to, "%Y-%m-%d").date()
                end_datetime = datetime.combine(end_date, time(23, 59, 59))
                transactions = transactions.filter(date__lte=end_datetime)
            except ValueError:
                transactions = transactions.filter(date__lte=date_to)
        if search:
            transactions = transactions.filter(
                Q(purpose__icontains=search) | Q(reference_number__icontains=search)
            )
        if verified_by and verified_by != "all":
            transactions = transactions.filter(verified_by=verified_by)

        # Apply pagination
        paginator = TransactionPagination()
        paginated_transactions = paginator.paginate_queryset(transactions, request)
        
        if paginated_transactions is not None:
            serializer = TransactionSerializer(paginated_transactions, many=True)
            return paginator.get_paginated_response(serializer.data)
        
        # Fallback if pagination fails
        serializer = TransactionSerializer(transactions, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=["get"])
    def summary(self, request, pk=None):
        """Get account summary with totals"""
        account = self.get_object()
        verified_transactions = account.transactions.filter(status="verified")

        agg = verified_transactions.aggregate(
            total_credits=Sum("amount", filter=Q(type="credit")),
            total_debits=Sum("amount", filter=Q(type="debit")),
        )
        total_credits = agg["total_credits"] or 0
        total_debits = agg["total_debits"] or 0

        return Response(
            {
                "account_id": account.id,
                "account_name": account.name,
                "current_balance": account.balance,
                "total_credits": total_credits,
                "total_debits": total_debits,
                "transaction_count": verified_transactions.count(),
                "created_at": account.created_at,
            }
        )


class TransactionViewSet(viewsets.ModelViewSet):
    # Staff logins are held to these; owners are unrestricted.
    required_permissions = {
        "GET": "banking.view",
        "POST": "banking.transact",
        "PUT": "banking.transact",
        "PATCH": "banking.transact",
        "DELETE": "banking.transact",
    }

    permission_classes = [IsAuthenticated, HasPermission]
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = ["account", "type", "nature", "category", "status", "verified_by"]
    search_fields = ["purpose", "reference_number"]
    ordering_fields = ["date", "amount", "type"]
    ordering = ["-date"]

    def get_queryset(self):
        """Return transactions based on user permissions"""
        user = owner_for(self.request)
        # If user is staff/admin, they can see all transactions
        if user.is_staff or user.is_superuser:
            return Transaction.objects.all()

        # Regular users can only see transactions for their own accounts
        return Transaction.objects.filter(account__owner=user)

    def get_serializer_class(self):
        if self.action in ["create", "update", "partial_update"]:
            return TransactionCreateSerializer
        return TransactionSerializer

    def perform_create(self, serializer):
        """Override to validate account ownership"""
        account = serializer.validated_data["account"]

        # Check if user owns the account (unless they're staff/admin)
        if not (self.request.user.is_staff or self.request.user.is_superuser):
            if account.owner != self.request.user:
                raise PermissionDenied(
                    "You can only create transactions for your own accounts"
                )

        # Save the transaction - balance will be updated in the model's save method
        serializer.save()

    @action(detail=False, methods=["get"])
    def my_transactions(self, request):
        """Get current user's transactions only"""
        transactions = Transaction.objects.filter(account__owner=owner_for(request))

        # Apply same filtering as regular endpoint
        transaction_type = request.query_params.get("type")
        status_filter = request.query_params.get("status")
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")
        search = request.query_params.get("search")
        account_id = request.query_params.get("account_id")
        verified_by = request.query_params.get("verified_by")

        if transaction_type and transaction_type != "all":
            transactions = transactions.filter(type=transaction_type)
        if status_filter and status_filter != "all":
            transactions = transactions.filter(status=status_filter)
        if date_from:
            transactions = transactions.filter(date__gte=date_from)
        if date_to:
            # Include the entire end date by adding time 23:59:59
            from datetime import datetime, time

            try:
                # Parse the date string and combine with end of day time
                end_date = datetime.strptime(date_to, "%Y-%m-%d").date()
                end_datetime = datetime.combine(end_date, time(23, 59, 59))
                transactions = transactions.filter(date__lte=end_datetime)
            except ValueError:
                # Fallback to original behavior if date parsing fails
                transactions = transactions.filter(date__lte=date_to)
        if search:
            transactions = transactions.filter(
                Q(purpose__icontains=search) | Q(reference_number__icontains=search)
            )
        if account_id:
            transactions = transactions.filter(account_id=account_id)
        if verified_by and verified_by != "all":
            transactions = transactions.filter(verified_by=verified_by)

        serializer = TransactionSerializer(transactions, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["get"])
    def employees(self, request):
        """Get list of employees who can verify transactions"""
        # Filter employees by current user (store owner)
        employees = (
            Employee.objects.filter(user=owner_for(request), status="active")
            .select_related("user")
            .order_by("name")
        )

        # Add search functionality
        search = request.query_params.get("search", "")
        if search:
            employees = employees.filter(
                Q(name__icontains=search)
                | Q(employee_id__icontains=search)
                | Q(role__icontains=search)
                | Q(department__icontains=search)
            )

        # Convert to the expected format for the frontend
        employee_data = [
            {
                "id": emp.id,  # Use Employee ID for verified_by field
                "first_name": emp.name.split()[0] if emp.name else "",
                "last_name": " ".join(emp.name.split()[1:])
                if len(emp.name.split()) > 1
                else "",
                "username": emp.employee_id,
                "email": emp.email,
                "full_name": emp.name,
                "name": emp.name,  # Add name field for consistency
                "employee_id": emp.employee_id,
                "role": emp.role,
                "department": emp.department,
                "status": emp.status,
            }
            for emp in employees
        ]

        return Response(employee_data)

    @action(detail=True, methods=["post"])
    def verify(self, request, pk=None):
        """Verify a pending transaction"""
        transaction = self.get_object()

        if transaction.status != "pending":
            return Response(
                {"error": "Transaction is not in pending status"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        transaction.status = "verified"
        # Get the employee associated with the current user
        try:
            employee = Employee.objects.get(user=owner_for(request), status="active")
            transaction.verified_by = employee
        except Employee.DoesNotExist:
            return Response(
                {"error": "No active employee found for current user"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        transaction.save()  # Balance will be updated in the model's save method

        serializer = self.get_serializer(transaction)
        return Response(serializer.data)

    @action(detail=True, methods=["post"])
    def cancel(self, request, pk=None):
        """Cancel a transaction"""
        transaction = self.get_object()

        if transaction.status == "cancelled":
            return Response(
                {"error": "Transaction is already cancelled"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # If transaction was verified, reverse the balance update
        if transaction.status == "verified":
            reverse_type = "debit" if transaction.type == "credit" else "credit"
            transaction.account.update_balance(transaction.amount, reverse_type)

        transaction.status = "cancelled"
        transaction.save()

        serializer = self.get_serializer(transaction)
        return Response(serializer.data)

    @action(detail=False, methods=["get"])
    def export_xlsx(self, request):
        """Export transactions to XLSX file"""
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        import io
        
        # Get filtered transactions using the same logic as other endpoints
        transactions = self.get_queryset()
        
        # Apply filtering
        transaction_type = request.query_params.get("type")
        status_filter = request.query_params.get("status")
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")
        search = request.query_params.get("search")
        account_id = request.query_params.get("account_id")
        verified_by = request.query_params.get("verified_by")

        if transaction_type and transaction_type != "all":
            transactions = transactions.filter(type=transaction_type)
        if status_filter and status_filter != "all":
            transactions = transactions.filter(status=status_filter)
        if date_from:
            transactions = transactions.filter(date__gte=date_from)
        if date_to:
            from datetime import datetime, time
            try:
                end_date = datetime.strptime(date_to, "%Y-%m-%d").date()
                end_datetime = datetime.combine(end_date, time(23, 59, 59))
                transactions = transactions.filter(date__lte=end_datetime)
            except ValueError:
                transactions = transactions.filter(date__lte=date_to)
        if search:
            transactions = transactions.filter(
                Q(purpose__icontains=search) | Q(reference_number__icontains=search)
            )
        if account_id:
            transactions = transactions.filter(account_id=account_id)
        if verified_by and verified_by != "all":
            transactions = transactions.filter(verified_by=verified_by)

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Transaction Statement"
        
        # Define header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = [
            "Date", "Reference Number", "Type", "Amount", "Purpose", 
            "Status", "Verified By", "Account", "Running Balance"
        ]
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add data rows with running balance calculation
        # To get the correct starting balance for the filtered date range,
        # compute the balance from all verified transactions BEFORE date_from.
        starting_balance = 0
        if account_id:
            try:
                account_obj = BankAccount.objects.get(id=account_id)
                # Start from the account's full balance and subtract all
                # verified transactions that are NOT in the filtered set.
                all_verified = account_obj.transactions.filter(status="verified")
                filtered_ids = set(transactions.filter(status="verified").values_list("id", flat=True))
                pre_agg = all_verified.exclude(id__in=filtered_ids).aggregate(
                    credits=Sum("amount", filter=Q(type="credit")),
                    debits=Sum("amount", filter=Q(type="debit")),
                )
                starting_balance = float(pre_agg["credits"] or 0) - float(pre_agg["debits"] or 0)
            except BankAccount.DoesNotExist:
                pass

        # Sort transactions by date (oldest first) for proper running balance
        sorted_transactions = transactions.order_by('date')
        running_balance = starting_balance

        # Now add rows with forward calculation
        row = 2
        for transaction in sorted_transactions:
            if transaction.status == "verified":
                if transaction.type == "credit":
                    running_balance += float(transaction.amount)
                else:
                    running_balance -= float(transaction.amount)
            
            ws.cell(row=row, column=1, value=transaction.date.strftime("%Y-%m-%d %H:%M:%S"))
            ws.cell(row=row, column=2, value=transaction.reference_number)
            ws.cell(row=row, column=3, value=transaction.get_type_display())
            ws.cell(row=row, column=4, value=float(transaction.amount))
            ws.cell(row=row, column=5, value=transaction.purpose)
            ws.cell(row=row, column=6, value=transaction.get_status_display())
            ws.cell(row=row, column=7, value=transaction.verified_by.name if transaction.verified_by else "N/A")
            ws.cell(row=row, column=8, value=transaction.account.name)
            ws.cell(row=row, column=9, value=running_balance if transaction.status == "verified" else "N/A")
            row += 1
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col)
            for cell in ws[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Generate filename with date range if specified
        filename = "transaction_statement"
        if date_from and date_to:
            filename += f"_{date_from}_to_{date_to}"
        elif date_from:
            filename += f"_from_{date_from}"
        elif date_to:
            filename += f"_to_{date_to}"
        filename += ".xlsx"
        
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @action(detail=False, methods=["get"])
    def dashboard_stats(self, request):
        """Get dashboard statistics for transactions"""
        account_id = request.query_params.get("account_id")

        transactions = self.get_queryset()
        if account_id:
            transactions = transactions.filter(account_id=account_id)

        verified_transactions = transactions.filter(status="verified")

        agg = verified_transactions.aggregate(
            total_credits=Sum("amount", filter=Q(type="credit")),
            total_debits=Sum("amount", filter=Q(type="debit")),
        )
        total_credits = agg["total_credits"] or 0
        total_debits = agg["total_debits"] or 0

        return Response(
            {
                "total_transactions": transactions.count(),
                "verified_transactions": verified_transactions.count(),
                "pending_transactions": transactions.filter(status="pending").count(),
                "total_credits": total_credits,
                "total_debits": total_debits,
                "net_amount": total_credits - total_debits,
            }
        )


class BankingPlanListView(generics.ListAPIView):
    """List all active banking plans"""

    queryset = BankingPlan.objects.filter(is_active=True)
    serializer_class = BankingPlanSerializer
    permission_classes = [IsAuthenticated, HasPermission]


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def activate_banking_plan(request):
    """Activate a banking plan for a user's account after successful payment"""
    account_id = request.data.get("account_id")
    plan_id = request.data.get("plan_id")
    payment_order_id = request.data.get("payment_order_id")
    payment_amount = request.data.get("payment_amount")

    if not all([account_id, plan_id, payment_order_id]):
        return Response(
            {
                "success": False,
                "message": "Account ID, Plan ID, and Payment Order ID are required",
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        # First get the banking plan
        plan = BankingPlan.objects.get(id=plan_id, is_active=True)

        # Try to get the specific account first
        account = None
        try:
            account = BankAccount.objects.get(id=account_id, owner=owner_for(request))
        except BankAccount.DoesNotExist:
            # Check if user has any existing accounts
            user_accounts = BankAccount.objects.filter(owner=owner_for(request))
            if user_accounts.exists():
                account = user_accounts.first()
            else:
                account_name = f"{request.user.first_name or request.user.username}'s Banking Account"
                account = BankAccount.objects.create(
                    owner=owner_for(request), name=account_name, balance=0.00
                )

        # Calculate expiry date based on plan period
        activated_at = timezone.now()
        if plan.period == "monthly":
            expires_at = activated_at + timedelta(days=30)
        elif plan.period == "yearly":
            expires_at = activated_at + timedelta(days=365)
        else:
            expires_at = None

        # Create or update user banking plan
        user_plan, created = UserBankingPlan.objects.get_or_create(
            user=owner_for(request),
            account=account,
            defaults={
                "plan": plan,
                "activated_at": activated_at,
                "expires_at": expires_at,
                "payment_order_id": payment_order_id,
                "payment_amount": payment_amount,
                "payment_status": "completed",
                "is_active": True,
            },
        )

        if not created:
            # Update existing plan
            user_plan.plan = plan
            user_plan.activated_at = activated_at
            user_plan.expires_at = expires_at
            user_plan.payment_order_id = payment_order_id
            user_plan.payment_amount = payment_amount
            user_plan.payment_status = "completed"
            user_plan.is_active = True
            user_plan.save()

        return Response(
            {
                "success": True,
                "message": f"Successfully activated {plan.name} {plan.period} plan for account {account.name}",
                "plan": BankingPlanSerializer(plan).data,
                "expires_at": expires_at,
            },
            status=status.HTTP_200_OK,
        )

    except BankAccount.DoesNotExist:
        return Response(
            {
                "success": False,
                "message": "Bank account not found or you do not have permission to access it",
            },
            status=status.HTTP_404_NOT_FOUND,
        )
    except BankingPlan.DoesNotExist:
        return Response(
            {"success": False, "message": "Banking plan not found"},
            status=status.HTTP_404_NOT_FOUND,
        )
    except Exception as e:
        return Response(
            {"success": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


class UserBankingPlanView(generics.RetrieveAPIView):
    """
    Get current user's banking plan
    """

    permission_classes = [IsAuthenticated, HasPermission]

    def get(self, request):
        try:
            user_plan = UserBankingPlan.objects.get(user=owner_for(request))
            from .serializers import UserBankingPlanSerializer

            return Response(UserBankingPlanSerializer(user_plan).data)
        except UserBankingPlan.DoesNotExist:
            return Response(
                {"message": "No active banking plan found"},
                status=status.HTTP_404_NOT_FOUND,
            )


@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def test_banking_endpoint(request):
    """Simple test endpoint to verify URL routing is working"""
    return Response(
        {
            "message": "Banking test endpoint is working",
            "method": request.method,
            "user": str(request.user),
        },
        status=status.HTTP_200_OK,
    )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def activate_banking_plan_v2(request):
    """Alternative activate banking plan function for testing"""
    return Response(
        {
            "success": True,
            "message": "Banking plan activation endpoint is working (v2)",
            "data": request.data,
        },
        status=status.HTTP_200_OK,
    )


class LoanViewSet(viewsets.ModelViewSet):
    """Loans the shop is repaying, and the installments paid against them."""
    # Staff logins are held to these; owners are unrestricted.
    required_permissions = {
        "GET": "banking.view",
        "POST": "banking.loans",
        "PUT": "banking.loans",
        "PATCH": "banking.loans",
        "DELETE": "banking.loans",
    }


    permission_classes = [IsAuthenticated, HasPermission]
    serializer_class = LoanSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ["status", "account"]
    search_fields = ["lender", "purpose"]
    ordering_fields = ["created_at", "start_date", "installment_amount"]
    ordering = ["-created_at"]

    def get_queryset(self):
        return (
            Loan.objects.filter(user=owner_for(self.request))
            .select_related("account")
            .prefetch_related("payments")
        )

    def perform_create(self, serializer):
        serializer.save(user=owner_for(self.request))

    @action(detail=False, methods=["get"])
    def summary(self, request):
        """Totals for the loan header, plus what falls due this month.

        `monthly_due` is the figure analytics treats as a fixed running cost —
        it leaves the business whether or not anything sells.
        """
        loans = self.get_queryset().filter(status="active")
        monthly = sum((loan.installment_amount for loan in loans), Decimal("0"))
        outstanding = sum((loan.remaining_amount for loan in loans), Decimal("0"))
        overdue = [loan for loan in loans if loan.is_overdue]
        return Response(
            {
                "active_count": loans.count(),
                "monthly_due": monthly,
                "outstanding": outstanding,
                "overdue_count": len(overdue),
                "overdue_amount": sum(
                    (loan.installment_amount for loan in overdue), Decimal("0")
                ),
                "next_due": min(
                    (loan.next_due_date for loan in loans if loan.next_due_date),
                    default=None,
                ),
            }
        )

    @action(
        detail=True,
        methods=["post", "delete"],
        url_path=r"payments/(?P<payment_id>\d+)/receipt",
        parser_classes=[MultiPartParser, FormParser],
    )
    def upload_receipt(self, request, pk=None, payment_id=None):
        """Attach the money receipt to an installment already paid, or drop it."""
        loan = self.get_object()
        try:
            payment = loan.payments.get(id=payment_id)
        except LoanPayment.DoesNotExist:
            return Response(
                {"error": "এই কিস্তিটা পাওয়া যায়নি।"}, status=status.HTTP_404_NOT_FOUND
            )

        if request.method == "DELETE":
            # Clear the field first so a missing file on disk still detaches
            # the row instead of leaving a link that 404s.
            if payment.receipt:
                payment.receipt.delete(save=False)
            payment.receipt = None
            payment.save(update_fields=["receipt"])
            return Response(
                LoanSerializer(
                    self.get_queryset().get(pk=loan.pk), context={"request": request}
                ).data
            )

        receipt = request.FILES.get("receipt")
        if receipt is None:
            return Response(
                {"error": "ফাইল পাওয়া যায়নি।"}, status=status.HTTP_400_BAD_REQUEST
            )
        problem = document_error(receipt)
        if problem:
            return Response({"error": problem}, status=status.HTTP_400_BAD_REQUEST)
        # Replacing a receipt must not leave the old file orphaned in MEDIA_ROOT.
        if payment.receipt:
            payment.receipt.delete(save=False)
        payment.receipt = receipt
        payment.save(update_fields=["receipt"])
        return Response(
            LoanSerializer(
                self.get_queryset().get(pk=loan.pk), context={"request": request}
            ).data
        )

    @action(
        detail=True,
        methods=["delete"],
        url_path=r"payments/(?P<payment_id>\d+)",
    )
    def remove_payment(self, request, pk=None, payment_id=None):
        """Undo an installment entered by mistake.

        LoanPayment.delete() also removes the bank transaction and credits the
        balance back, so the books do not keep a payment that never happened.
        A closed loan reopens, since it is no longer fully repaid.
        """
        loan = self.get_object()
        try:
            payment = loan.payments.get(id=payment_id)
        except LoanPayment.DoesNotExist:
            return Response(
                {"error": "এই কিস্তিটা পাওয়া যায়নি।"}, status=status.HTTP_404_NOT_FOUND
            )

        with db_transaction.atomic():
            payment.delete()
            if loan.status == "closed":
                loan.status = "active"
                loan.save(update_fields=["status"])

        return Response(
            {
                "message": "কিস্তিটা বাতিল হয়েছে।",
                "loan": LoanSerializer(
                    self.get_queryset().get(pk=loan.pk), context={"request": request}
                ).data,
            }
        )

    @action(detail=True, methods=["post"])
    def pay(self, request, pk=None):
        """Record an installment.

        Writes the bank transaction too, so the money is not counted twice:
        analytics reads loan cost from LoanPayment and skips any transaction
        already attached to one.
        """
        loan = self.get_object()
        if loan.status != "active":
            return Response(
                {"error": "এই লোনটা আর চালু নেই।"}, status=status.HTTP_400_BAD_REQUEST
            )

        try:
            amount = Decimal(str(request.data.get("amount") or loan.installment_amount))
        except (InvalidOperation, TypeError):
            return Response(
                {"error": "টাকার অঙ্কটা ঠিক নেই।"}, status=status.HTTP_400_BAD_REQUEST
            )
        if amount <= 0:
            return Response(
                {"error": "টাকার পরিমাণ শূন্যের বেশি হতে হবে।"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        paid_on = request.data.get("paid_on") or timezone.localdate().isoformat()
        reference = request.data.get("reference") or ""

        with db_transaction.atomic():
            txn = None
            if loan.account_id:
                txn = Transaction.objects.create(
                    account=loan.account,
                    type="debit",
                    nature="payment",
                    category="other",
                    amount=amount,
                    purpose=f"লোনের কিস্তি — {loan.lender}",
                    status="verified",
                )
            payment = LoanPayment.objects.create(
                loan=loan,
                amount=amount,
                paid_on=paid_on,
                transaction=txn,
                reference=reference,
            )
            # Close the loan once it is fully repaid, so it drops out of the
            # monthly cost projection instead of lingering forever.
            if loan.remaining_amount <= 0 or loan.paid_count >= loan.installment_count:
                loan.status = "closed"
                loan.save(update_fields=["status"])

        return Response(
            {
                "message": "কিস্তি জমা হয়েছে।",
                "payment": LoanPaymentSerializer(payment).data,
                "loan": LoanSerializer(
                    self.get_queryset().get(pk=loan.pk), context={"request": request}
                ).data,
            },
            status=status.HTTP_201_CREATED,
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def expense_categories(request):
    """Suggestions for the খাত picker: the presets plus whatever this shop
    has already typed, so a custom category only has to be spelled once."""
    used = (
        Transaction.objects.filter(account__owner=owner_for(request))
        .exclude(category="")
        .values_list("category", flat=True)
        .distinct()
    )
    presets = [key for key, _ in Transaction.CATEGORY_CHOICES]
    extra = sorted(set(used) - set(presets))
    return Response({"presets": presets, "custom": extra})


class RecurringCostViewSet(viewsets.ModelViewSet):
    """Fixed monthly bills — office rent and the like."""
    # Staff logins are held to these; owners are unrestricted.
    required_permissions = {
        "GET": "banking.view",
        "POST": "banking.costs",
        "PUT": "banking.costs",
        "PATCH": "banking.costs",
        "DELETE": "banking.costs",
    }


    permission_classes = [IsAuthenticated, HasPermission]
    serializer_class = RecurringCostSerializer
    # A receipt upload is multipart; the rest of the viewset stays JSON.
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ["is_active", "category"]
    search_fields = ["title", "notes"]

    def get_queryset(self):
        return (
            RecurringCost.objects.filter(user=owner_for(self.request))
            .select_related("account")
            .prefetch_related("payments")
        )

    def perform_create(self, serializer):
        serializer.save(user=owner_for(self.request))

    @action(detail=False, methods=["get"])
    def summary(self, request):
        costs = self.get_queryset().filter(is_active=True)
        overdue = [cost for cost in costs if cost.is_overdue]
        unpaid = [cost for cost in costs if not cost.paid_this_month]
        return Response(
            {
                "active_count": costs.count(),
                "monthly_total": sum((c.amount for c in costs), Decimal("0")),
                "unpaid_count": len(unpaid),
                "unpaid_amount": sum((c.amount for c in unpaid), Decimal("0")),
                "overdue_count": len(overdue),
            }
        )

    @action(detail=True, methods=["post"])
    def pay(self, request, pk=None):
        """Settle one month.

        The period defaults to the current month; `unique_together` on
        (cost, period) is what stops the same month being paid twice.
        """
        cost = self.get_object()
        period_raw = request.data.get("period")
        period = (
            date.fromisoformat(period_raw) if period_raw else cost.current_period
        ).replace(day=1)

        if cost.payments.filter(period=period).exists():
            return Response(
                {"error": "এই মাসের টাকা আগেই দেওয়া হয়েছে।"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            amount = Decimal(str(request.data.get("amount") or cost.amount))
        except (InvalidOperation, TypeError):
            return Response(
                {"error": "টাকার অঙ্কটা ঠিক নেই।"}, status=status.HTTP_400_BAD_REQUEST
            )

        with db_transaction.atomic():
            txn = None
            if cost.account_id:
                txn = Transaction.objects.create(
                    account=cost.account,
                    type="debit",
                    nature="expense",
                    category=cost.category,
                    amount=amount,
                    purpose=f"{cost.title} — {period:%B %Y}",
                    status="verified",
                )
            payment = RecurringCostPayment.objects.create(
                cost=cost, period=period, amount=amount, transaction=txn
            )

        return Response(
            {
                "message": "টাকা জমা হয়েছে।",
                "payment": RecurringCostPaymentSerializer(
                    payment, context={"request": request}
                ).data,
                "cost": RecurringCostSerializer(self.get_queryset().get(pk=cost.pk)).data,
            },
            status=status.HTTP_201_CREATED,
        )

    @action(
        detail=True,
        methods=["post", "delete"],
        url_path=r"payments/(?P<payment_id>\d+)/receipt",
        parser_classes=[MultiPartParser, FormParser],
    )
    def upload_receipt(self, request, pk=None, payment_id=None):
        """Attach the money receipt to a month already paid, or drop it."""
        cost = self.get_object()
        try:
            payment = cost.payments.get(id=payment_id)
        except RecurringCostPayment.DoesNotExist:
            return Response(
                {"error": "এই মাসের রেকর্ড পাওয়া যায়নি।"},
                status=status.HTTP_404_NOT_FOUND,
            )

        if request.method == "DELETE":
            if payment.receipt:
                payment.receipt.delete(save=False)
            payment.receipt = None
            payment.save(update_fields=["receipt"])
            return Response(
                RecurringCostPaymentSerializer(
                    payment, context={"request": request}
                ).data
            )
        receipt = request.FILES.get("receipt")
        if receipt is None:
            return Response(
                {"error": "ফাইল পাওয়া যায়নি।"}, status=status.HTTP_400_BAD_REQUEST
            )
        problem = document_error(receipt)
        if problem:
            return Response({"error": problem}, status=status.HTTP_400_BAD_REQUEST)
        # Replacing a receipt must not leave the old file orphaned in MEDIA_ROOT.
        if payment.receipt:
            payment.receipt.delete(save=False)
        payment.receipt = receipt
        payment.save(update_fields=["receipt"])
        return Response(
            RecurringCostPaymentSerializer(payment, context={"request": request}).data
        )

    @action(
        detail=True,
        methods=["delete"],
        url_path=r"payments/(?P<payment_id>\d+)",
    )
    def remove_payment(self, request, pk=None, payment_id=None):
        """Undo a month entered by mistake; the bank transaction goes too."""
        cost = self.get_object()
        try:
            payment = cost.payments.get(id=payment_id)
        except RecurringCostPayment.DoesNotExist:
            return Response(
                {"error": "এই মাসের রেকর্ড পাওয়া যায়নি।"},
                status=status.HTTP_404_NOT_FOUND,
            )
        payment.delete()
        return Response(
            {
                "message": "বাতিল হয়েছে।",
                "cost": RecurringCostSerializer(self.get_queryset().get(pk=cost.pk)).data,
            }
        )
